import command_system
import vk
#from settings import s_token, token
from openpyxl import load_workbook
import time

session = vk.Session()
api = vk.API(session, v=5.92)

def upcoming():
    #posts = api.wall.search(owner_id=-177456743, domain = 'test_bot_df', query = '#событие', count = 1, access_token = s_token)
    #message = posts['items'][0]['text']
    message=''
    unixtime = round(time.time()+10800)
    wb = load_workbook('/home/jBloodless/mysite/commands/schedule.xlsx', data_only=True)
    sheet = wb['Лист1']
    for i in range (2,45):
        if sheet.cell(row=i, column=5).value>=unixtime:
            j=i-1
            message=str(sheet.cell(row=j, column=3).value)
            attachment=str(sheet.cell(row=j, column=6).value)
            break
    #attachment=''
    buttons = open("/home/jBloodless/mysite/defaultKeys.json", "r", encoding="UTF-8").read()
    return message, attachment, buttons

upcoming_command = command_system.Command()

upcoming_command.keys = ['Ближайшее событие']
upcoming_command.description = 'Пришлю событие'
upcoming_command.process = upcoming

