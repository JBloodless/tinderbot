import command_system
import vk
from settings import token, s_token
from flask_app import curr_user_id
import random
import os
import re
import requests
from flask import json
from openpyxl import load_workbook
import time

session = vk.Session()
api = vk.API(session, v=5.92)

def tinderLike():
    user_id = curr_user_id()
    curr_user = api.users.get(user_ids=str(user_id), fields = 'sex', access_token = s_token)
    curr_sex = curr_user[0]['sex']
    first_name = curr_user[0]['first_name']
    last_name = curr_user[0]['last_name']
    if curr_sex == 1:
        sex_str = 'female'
    elif curr_sex == 2:
        sex_str = 'male'
    hist = api.messages.getHistory(peer_id = user_id, group_id=17379212, offset = 1, count = 1, access_token = token)
    tinderBio = hist['items'][0]['text']
    l = tinderBio.find('d')
    r = tinderBio.find('|')
    match_id = tinderBio[l+1:r]

    path = "/home/jBloodless/mysite/tinder/"
    dirs = os.listdir(path)
    for i in range(len(dirs)):
        if dirs[i].endswith('_{0}_{1}_{2}'.format(sex_str, first_name, last_name)):
            dir_name = re.split(r'_', str(dirs[i]))
            currdest = dir_name[0]
    match_user = api.users.get(user_ids=str(match_id), fields = 'sex', access_token = s_token)
    match_sex = match_user[0]['sex']
    if match_sex == 1:
        match_sex_str = 'female'
    elif match_sex == 2:
        match_sex_str = 'male'
    match_first_name = match_user[0]['first_name']
    match_last_name = match_user[0]['last_name']
    match_prof=match_first_name+'_'+match_last_name

    curr_like_f= open("/home/jBloodless/mysite/tinder/{0}_{1}_{2}_{3}/likes/{4}_{5}".format(currdest, sex_str, first_name, last_name, match_first_name, match_last_name), "w")
    curr_like_f.write(match_id)
    curr_like_f.close()

    path = "/home/jBloodless/mysite/tinder/"
    match_dirs = os.listdir(path)
    for i in range(len(match_dirs)):
        if match_dirs[i].endswith('_{0}_{1}_{2}'.format(match_sex_str, match_first_name, match_last_name)):
            match_dir_name = re.split(r'_', str(match_dirs[i]))
            matchdest = match_dir_name[0]
    if os.path.isfile("/home/jBloodless/mysite/tinder/{0}_{1}_{2}_{3}/likes/{4}_{5}".format(matchdest, match_sex_str, match_first_name, match_last_name, first_name, last_name)):

        unixtime = round(time.time()+10800)
        wb = load_workbook('/home/jBloodless/mysite/commands/schedule.xlsx', data_only=True)
        sheet = wb['Лист1']
        for i in range (2,45):
            if sheet.cell(row=i, column=5).value>=unixtime:
                j=i-1
                message_add=str(sheet.cell(row=j, column=3).value)
                attachment=str(sheet.cell(row=j, column=6).value)
                break
        #message = attachment
        message = 'Взаимный лайк! Совпадение или судьба?... Напиши по ссылке и выясни! Вам даже не стоит ломать голову над местом для свидания, ведь вы на ДФе!\n'+str(message_add)+ '\nЕсли не хватило, то можно начать заново:'
        message_match = 'Взаимный лайк c @id'+str(user_id)+'! Совпадение или судьба?... Напиши по ссылке и выясни! Вам даже не стоит ломать голову над местом для свидания, ведь вы на ДФе!\n'+str(message_add)+ '\nЕсли не хватило, то можно начать заново:'
        rnd_seed = random.randint(0, 2147483647)
        buttons = open("/home/jBloodless/mysite/tinderEnd.json", "r", encoding="UTF-8").read()
        api.messages.send(random_id=rnd_seed, access_token=token, peer_id=match_id, group_id = 17379212, message=message_match, attachment=attachment, keyboard=buttons)
        return message, attachment, buttons
    else:
        path = "/home/jBloodless/mysite/tinder/"
        user_id = curr_user_id()
        curr_user = api.users.get(user_ids=str(user_id), fields = 'sex', access_token = s_token)
        curr_sex = curr_user[0]['sex']
        if curr_sex == 1:
            curr_sex_str = 'female'
        elif curr_sex == 2:
            curr_sex_str = 'male'
        first_name = curr_user[0]['first_name']
        last_name = curr_user[0]['last_name']

        dirs = os.listdir(path)
        for i in range(len(dirs)):
            if dirs[i].endswith('_{0}_{1}_{2}'.format(curr_sex_str, first_name, last_name)):
                dir_name = re.split(r'_', str(dirs[i]))
                currdest = dir_name[0]

        user_name = first_name
        user_surn = last_name
        sex=''

        dislikes = '/home/jBloodless/mysite/tinder/{0}_{1}_{2}_{3}/dislikes'.format(currdest, curr_sex_str, first_name, last_name)
        likes = '/home/jBloodless/mysite/tinder/{0}_{1}_{2}_{3}/likes'.format(currdest, curr_sex_str, first_name, last_name)
        i=0
        if currdest=='toA':
            for i in range(len(os.listdir(path))):
                if i==len(os.listdir(path))-1:
                    message = 'Пока новых участников не появилось. Но не стоит отчаиваться — это только начало! Возвращайся позже, и чудо обязательно произойдет!'
                    buttons = open("/home/jBloodless/mysite/tinderEnd.json", "r", encoding="UTF-8").read()
                    return message, '', buttons
                name = re.split(r'_', str(os.listdir(path)[i]))
                dest = name[0]
                sex = name[1]
                user_name=name[2]
                user_surn=name[3]
                user_prof = user_name+'_'+user_surn
                if user_prof not in os.listdir(dislikes) and user_prof not in os.listdir(likes) and first_name!=user_name and last_name!=user_surn:
                    break

        elif currdest=='toM':
            for i in range(len(os.listdir(path))):
                if i==len(os.listdir(path))-1:
                    message = 'Пока новых участников не появилось. Но не стоит отчаиваться — это только начало! Возвращайся позже, и чудо обязательно произойдет!'
                    buttons = open("/home/jBloodless/mysite/tinderEnd.json", "r", encoding="UTF-8").read()
                    return message, '', buttons
                name = re.split(r'_', str(os.listdir(path)[i]))
                dest = name[0]
                sex = name[1]
                user_name=name[2]
                user_surn=name[3]
                user_prof = user_name+'_'+user_surn
                if user_prof not in os.listdir(dislikes) and user_prof not in os.listdir(likes) and first_name!=user_name and last_name!=user_surn and sex!='female':
                    break

        elif currdest=='toW':
            for i in range(len(os.listdir(path))):
                if i==len(os.listdir(path))-1:
                    message = 'Пока новых участников не появилось. Но не стоит отчаиваться — это только начало! Возвращайся позже, и чудо обязательно произойдет!'
                    buttons = open("/home/jBloodless/mysite/tinderEnd.json", "r", encoding="UTF-8").read()
                    return message, '', buttons
                name = re.split(r'_', str(os.listdir(path)[i]))
                dest = name[0]
                sex = name[1]
                user_name=name[2]
                user_surn=name[3]
                user_prof = user_name+'_'+user_surn
                if user_prof not in os.listdir(dislikes) and user_prof not in os.listdir(likes) and first_name!=user_name and last_name!=user_surn and sex!='male':
                    break

    #user_dir = "/home/jBloodless/mysite/tinder/{0}_{1}_{2}_{3}".format(name[0], sex, user_name, user_surn)
        user_pic = "/home/jBloodless/mysite/tinder/{0}_{1}_{2}_{3}/{2}_{3}.jpg".format(dest, sex, user_name, user_surn)
        user_inf = "/home/jBloodless/mysite/tinder/{0}_{1}_{2}_{3}/{2}_{3}.txt".format(dest, sex, user_name, user_surn)
        id_txt = "/home/jBloodless/mysite/tinder/{0}_{1}_{2}_{3}/user_id.txt".format(dest, sex, user_name, user_surn)

        bio = open(user_inf, "r", encoding="UTF-8")
        id_to_msg = open(id_txt, "r", encoding="UTF-8")
        message = "@id"+str(id_to_msg.read())+'\n'+str(bio.read())
    #message = currdest

        server = api.photos.getMessagesUploadServer(peer_id=curr_user_id(), access_token = token)
        upload_url = server["upload_url"]
        files = {'photo': open(user_pic, 'rb')}
        response = requests.post(upload_url, files=files)
        result = json.loads(response.text)
        uploadResult = api.photos.saveMessagesPhoto(server=result["server"], photo=result["photo"], hash=result["hash"], access_token = token)
        attachment = 'photo'+str(uploadResult[0]["owner_id"])+'_'+str(uploadResult[0]["id"])
    #attachment = ''

        buttons = open("/home/jBloodless/mysite/tinderKeys.json", "r", encoding="UTF-8").read()
    return message, attachment, buttons

tinderLike_command = command_system.Command()

tinderLike_command.keys = ['Лайк!']
tinderLike_command.description = 'internal tinder message'
tinderLike_command.process = tinderLike